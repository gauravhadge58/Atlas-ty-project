"""
material_db.py
==============
SQLite-backed persistence for the material reference table.

On first access, if the database does not exist, it is seeded automatically
from the bundled `data/material_reference.json`.

Schema (table: material_reference)
-----------------------------------
  id            INTEGER PRIMARY KEY AUTOINCREMENT
  edm           TEXT NOT NULL
  materials     TEXT NOT NULL   -- JSON array serialised as a string
  finish        TEXT NOT NULL   -- JSON array serialised as a string
  heat          TEXT            -- NULL when no heat-treatment validation required
  base_material TEXT            -- descriptive name from the Excel "Base Material" column (optional)
  use_notes     TEXT            -- free-text "Use" column from Excel (optional)
"""
from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional

_BASE_DIR = Path(__file__).resolve().parent.parent          # backend/
_DB_PATH   = _BASE_DIR / "data" / "material_reference.db"
_JSON_PATH = _BASE_DIR / "data" / "material_reference.json"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS material_reference (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            edm           TEXT NOT NULL,
            materials     TEXT NOT NULL,
            finish        TEXT NOT NULL,
            heat          TEXT,
            base_material TEXT,
            use_notes     TEXT
        )
        """
    )
    conn.commit()


def _seed_from_json(conn: sqlite3.Connection) -> None:
    """Populate DB from the bundled JSON file (run once on first startup)."""
    if not _JSON_PATH.exists():
        return
    with _JSON_PATH.open("r", encoding="utf-8") as f:
        rows: List[Dict[str, Any]] = json.load(f)
    conn.executemany(
        """
        INSERT INTO material_reference (edm, materials, finish, heat, base_material, use_notes)
        VALUES (:edm, :materials, :finish, :heat, :base_material, :use_notes)
        """,
        [
            {
                "edm":           row.get("edm", ""),
                "materials":     json.dumps(row.get("materials") or []),
                "finish":        json.dumps(row.get("finish") or []),
                "heat":          row.get("heat"),
                "base_material": row.get("base_material"),
                "use_notes":     row.get("use_notes"),
            }
            for row in rows
        ],
    )
    conn.commit()


def _init_db() -> None:
    """Create the DB + seed from JSON if this is a first run."""
    first_run = not _DB_PATH.exists()
    conn = _get_connection()
    try:
        _ensure_table(conn)
        if first_run:
            _seed_from_json(conn)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_reference_table() -> List[Dict[str, Any]]:
    """Return all rows as a list of dicts (materials/finish as Python lists)."""
    _init_db()
    conn = _get_connection()
    try:
        rows = conn.execute(
            "SELECT id, edm, materials, finish, heat, base_material, use_notes "
            "FROM material_reference ORDER BY id"
        ).fetchall()
        return [_row_to_dict(r) for r in rows]
    finally:
        conn.close()


def save_reference_table(rows: List[Dict[str, Any]]) -> None:
    """
    Replace the entire table with *rows*.
    Each row must have at minimum: edm, materials (list), finish (list).
    """
    _init_db()
    conn = _get_connection()
    try:
        conn.execute("DELETE FROM material_reference")
        conn.executemany(
            """
            INSERT INTO material_reference (edm, materials, finish, heat, base_material, use_notes)
            VALUES (:edm, :materials, :finish, :heat, :base_material, :use_notes)
            """,
            [
                {
                    "edm":           row.get("edm", ""),
                    "materials":     json.dumps(_to_list(row.get("materials"))),
                    "finish":        json.dumps(_to_list(row.get("finish"))),
                    "heat":          row.get("heat") or None,
                    "base_material": row.get("base_material"),
                    "use_notes":     row.get("use_notes"),
                }
                for row in rows
            ],
        )
        conn.commit()
    finally:
        conn.close()


def reset_to_default() -> None:
    """
    Drop all rows and re-seed from the bundled JSON so the user
    can recover the factory defaults at any time.
    """
    _init_db()
    conn = _get_connection()
    try:
        conn.execute("DELETE FROM material_reference")
        conn.commit()
        _seed_from_json(conn)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

_EDM_PATTERN = re.compile(r"\bEDM\d+", re.IGNORECASE)


def parse_excel_to_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parse an .xlsx file whose format matches the company material sheet:

        Sr. No. | Base Material | EDM/ Grade | Surface Finish | Heat Treatment | Use

    The "EDM/ Grade" cell contains the EDM code and material aliases separated
    by "/".  E.g.:  "EDM000136/ En8/ 80M40/SM45C N"

    Returns a list of dicts ready to be passed to save_reference_table().
    """
    import io
    import openpyxl  # imported lazily so the rest of the app works without it

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # ---------- find header row ----------
    header_row_idx: Optional[int] = None
    col_map: Dict[str, int] = {}   # canonical_name -> 0-based col index

    _ALIASES = {
        "sr":             ["sr. no.", "sr.no", "sr no", "no.", "#", "s.no"],
        "base_material":  ["base material", "material name", "base material\n"],
        "edm_grade":      ["edm/ grade", "edm/grade", "edm grade", "edm", "grade"],
        "surface_finish": ["surface finish", "finish", "surface finish\n"],
        "heat":           ["heat treatment", "heat", "heat treatment\n"],
        "use":            ["use", "uses", "application"],
    }

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        row_text = [str(c).strip().lower() if c is not None else "" for c in row]
        # Check if this row looks like a header (has "edm" somewhere)
        if any("edm" in cell for cell in row_text):
            header_row_idx = row_idx
            for col_idx, cell_text in enumerate(row_text):
                for canonical, aliases in _ALIASES.items():
                    if any(alias in cell_text for alias in aliases):
                        col_map.setdefault(canonical, col_idx)
            break

    if header_row_idx is None or "edm_grade" not in col_map:
        raise ValueError(
            "Could not locate the header row. "
            "Make sure the sheet has a column labelled 'EDM/ Grade'."
        )

    # ---------- parse data rows ----------
    results: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        def cell(key: str) -> str:
            idx = col_map.get(key)
            if idx is None:
                return ""
            val = row[idx]
            return str(val).strip() if val is not None else ""

        edm_grade_raw = cell("edm_grade")
        if not edm_grade_raw:
            continue  # skip blank rows

        # Split on "/"
        parts = [p.strip() for p in edm_grade_raw.split("/") if p.strip()]
        if not parts:
            continue

        # First part that matches EDM pattern is the code; rest are material names
        edm_code = ""
        material_parts: List[str] = []
        for i, p in enumerate(parts):
            if _EDM_PATTERN.match(p):
                edm_code = p
            else:
                material_parts.append(p)

        if not edm_code:
            # Fallback: first token is the code
            edm_code = parts[0]
            material_parts = parts[1:]

        # Surface finish — split by "/"
        finish_raw = cell("surface_finish")
        finish_list = [f.strip() for f in finish_raw.split("/") if f.strip()] if finish_raw else []

        # Heat treatment — keep full text; null if NA or empty
        heat_raw = cell("heat").strip()
        heat_val: Optional[str] = None
        if heat_raw and heat_raw.upper() not in ("NA", "N/A", "NONE", "-", ""):
            heat_val = heat_raw

        results.append(
            {
                "edm":           edm_code,
                "materials":     material_parts,
                "finish":        finish_list,
                "heat":          heat_val,
                "base_material": cell("base_material") or None,
                "use_notes":     cell("use") or None,
            }
        )

    if not results:
        raise ValueError("No data rows found in the uploaded Excel file.")

    return results


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _row_to_dict(row: sqlite3.Row) -> Dict[str, Any]:
    d = dict(row)
    d["materials"] = json.loads(d.get("materials") or "[]")
    d["finish"]    = json.loads(d.get("finish") or "[]")
    return d


def _to_list(value: Any) -> List[str]:
    if isinstance(value, list):
        return [str(v) for v in value if v]
    if isinstance(value, str):
        # Could be a JSON string or a plain comma-separated string
        stripped = value.strip()
        if stripped.startswith("["):
            try:
                return [str(v) for v in json.loads(stripped) if v]
            except json.JSONDecodeError:
                pass
        return [v.strip() for v in stripped.split(",") if v.strip()]
    return []
